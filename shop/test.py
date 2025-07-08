import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook

# Path to the Excel file and WebDriver
EXCEL_FILE = "test_cases.xlsx"
DRIVER_PATH = "C:\Program Files\Google\Chrome"

# URL of the page to test
URL = "http://localhost/QuickBite-Instant_Food_Availability_and_Preordering_Solution/index.php"

# Initialize WebDriver
driver = webdriver.Chrome(DRIVER_PATH)
driver.get(URL)
driver.maximize_window()

# Log in as a shop owner (if required)
def login():
    try:
        driver.find_element(By.ID, "username").send_keys("canteen")
        driver.find_element(By.ID, "password").send_keys("123")
        driver.find_element(By.ID, "loginButton").click()
        time.sleep(2)
    except Exception as e:
        print("Login not required or failed: ", e)

# Read test cases from Excel
workbook = load_workbook(EXCEL_FILE)
sheet = workbook.active

# Iterate through test cases
for row in range(2, sheet.max_row + 1):
    menu_name = sheet.cell(row=row, column=1).value
    price = sheet.cell(row=row, column=2).value
    quantity = sheet.cell(row=row, column=3).value
    expected_result = sheet.cell(row=row, column=4).value

    try:
        # Navigate to the "Add Menu" page
        driver.find_element(By.LINK_TEXT, "Add Menu").click()
        time.sleep(1)

        # Fill in the form
        driver.find_element(By.ID, "menuName").send_keys(menu_name)
        driver.find_element(By.ID, "menuPrice").send_keys(price)
        driver.find_element(By.ID, "menuQuantity").send_keys(quantity)
        driver.find_element(By.ID, "addMenuButton").click()
        time.sleep(2)

        # Check for success or failure notification
        success_message = "Successfully add new menu."
        failure_message = "Failed to add menu."

        if success_message in driver.page_source:
            actual_result = "Success"
        elif failure_message in driver.page_source:
            actual_result = "Failure"
        else:
            actual_result = "Unknown"

        # Write the result back to the Excel sheet
        sheet.cell(row=row, column=5).value = actual_result

    except Exception as e:
        print(f"Test case {row - 1} failed: {e}")
        sheet.cell(row=row, column=5).value = "Error"

    # Navigate back to the menu list
    driver.get(URL)
    time.sleep(2)

# Save the updated Excel file
workbook.save(EXCEL_FILE)

# Close the WebDriver
driver.quit()

print("Test execution completed. Results updated in the Excel file.")
